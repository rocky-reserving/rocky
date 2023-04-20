"""
This module contains the `ROCKY3` class, which is the main class for the `ROCKY3` package.
The `ROCKY3` modelling framework is meant to be a flexible and extensible framework for
reserve modelling on loss triangles.

The `ROCKY3` model object is a container that holds:
    1. A dictionary of triangles inheriting from the  `Triangle` object, that is itself a flexible
       container holding multiple types of triangular reserve analysis data (e.g., paid loss,
       reported loss, counts, etc.) as well as exposure data and selections such as  expected
       loss ratios, selected development factors, etc. Will allow for only a single triangle 
       from each of the following categories:
        - Paid loss triangles ("Paid Losses")
        - Paid DCCE triangles ("Paid DCCE")
        - Reported loss triangles ("Reported Losses")
        - Reported claim count triangles ("Reported Counts")
        - Closed claim count triangles
       From these, several other triangles can be derived, including:
        - Case reserves
        - Open claim counts
        - Paid / reported losses (or reported / paid losses)
        - Closed / reported counts (or reported / closed counts)
        - Average reported losses per claim
        - Average paid losses per claim
        - Average case outstanding reserves per claim
    2. Various models of different types -- started with `Poisson` and `Gamma` GLMs + standard
       `Mack`, but will be expanded to include other models (e.g., `Ridge`, `Lasso`, etc.)
       using class inheritance to assign models to the `ROCKY3` class
        - These models are stored as a dictionary of model objects that attributes of the `ROCKY3`
          class. Each model object has the following attributes:
          - `name`: 
              the name of the model (e.g., 'mack', 'odpGLM', etc.) that is used
              to identify the model in the `ROCKY3` class. This also allows for multiple models
              of the same type to be stored in the `ROCKY3` class (e.g., "mack_vwa_all",
              "mack_medial_exhi", etc.).
          - `model`:
              the model object (e.g., a `scikit-learn` model object customized for this
              package)
          - `model_type`:
              the class the model is an instance of (e.g., `Poisson`, `Gamma`, etc.)
        - Includes a `base` model for each type that is all accident/development parameters and
          calendar period parameters set to 0. This allows for the `base` model to be fit to the
          data, and requires that all refinements to the model be done by combining parameters
          from the base model, and considering whether the loss of accuracy is worth the gain
          in gerenalizability. This is done automatically whenever assigning a model to the
          `ROCKY3` class.
        - The base model is also regularized by default, so that the model is not necessarily
          overfit to the data. Thus any adjustments to the base model should show a significant
          improvement in the model's performance.
        - Each new model type will be a subclass of the base model type, and will inherit all
          of the base model's methods and attributes. The new model type will then add new
          methods and attributes that are specific to the new model type.
        - Each model will have a 'doc' method that will print an html page with the model's
          documentation, including a detailed mathematical description of the model, a list of
          the model's parameters, and a list of the model's methods.
    3. A validation object that contains the results of the model's validation and diagnostic 
       testing. This includes:
         - The model's performance metrics (e.g., RMSE, MAE, R^2, etc.)
         - The model's feature importances
         - The model's residuals
         - The model's standardized residuals
    4. A forecasting object that contains assumptions about the future (e.g., future development
       factors, future loss ratios, etc.) and can be used to forecast future loss emergengence for
       each model/loss type.
    5. A plotting object that is used to build interactive plots of the data and models. Uses the 
       plotly library to have the interactivity. Has standard reserve analysis plots:
        - Percent of loss emergence by development period
        - Loss ratios by development period
        - Standardized residual plots by development period, accident period, calendar period, etc.
        - Actual vs. predicted loss emergence by development period
        - qq vs pp plots 

"""
import numpy as np
import pandas as pd
import re

from forecast import Forecast
from triangle import Triangle
from glm import Poisson, Gamma
from chainladder import ChainLadder

from dataclasses import dataclass
from typing import Any, Optional

from tkinter import filedialog
from openpyxl import load_workbook
from warnings import filterwarnings
filterwarnings("ignore", category=UserWarning, module="openpyxl")


# add all implemented model types to this list
all_models = 'Poisson Gamma'.split()

from dataclasses import dataclass, is_dataclass

@dataclass
class rockyObj:
    id: str = None
    obj: object = None

@dataclass
class rockyContainer:
    def __repr__(self) -> str:
        # only show comma-separated list of ids if there are any objects in the container
        if len(self.__dict__) > 0:
            return ', '.join([f'"{k}"' for k in self.__dict__.keys()])
        else:
            return '()'
        
    def add(self, _obj) -> None:
        setattr(self, _obj.id, _obj)

@dataclass
class ROCKY:
    # initialize the model attribute, triangle attribute, forecast attribute,
    # validation attribute, and plotting attribute
    id: str = None
    model: Any = rockyContainer()
    tri: Any = rockyContainer()
    forecast: Any = rockyContainer()
    validate: Any = rockyContainer()

    def set_id(self, id: str) -> None:
        """
        Set the id of the model.
        """
        # check that the id is a string
        if not isinstance(id, str):
            raise TypeError('The id must be a string.')
        
        # check that the id is unique
        if id in self.__dict__.keys():
            raise ValueError(f'The id "{id}" is not unique.')
        
        # set the id
        self.id = id

    def tri_from_clipboard(self,
                           id: str,
                           origin_columns: int = 1
                           ) -> None:
        """
        Create a triangle from the clipboard. Calls the `from_clipboard` method of the
        `Triangle` class.

        Parameters
        ----------
        id : str
            The id of the triangle.
        origin_columns : int, optional
            The number of origin columns in the triangle. The default is 1,
            which means that the first column is the the origin (eg accident 
            year) and the remaining columns are the development periods.

        Returns
        -------
        No return value. The triangle is added to the model.
        """
        # create the triangle
        tri = Triangle.from_clipboard(id=id, origin_columns=origin_columns)

        # add the triangle to the model
        self.add_triangle(tri)

    def tri_from_csv(self,
                     id: str,
                     file_path: str,
                     origin_columns: int = 1
                     ) -> None:
        """
        Create a triangle from a csv file. Calls the `from_csv` method of the
        `Triangle` class.

        Parameters
        ----------
        id : str
            The id of the triangle.
        file_path : str
            The path to the csv file.
        origin_columns : int, optional
            The number of origin columns in the triangle. The default is 1,
            which means that the first column is the the origin (eg accident
            year) and the remaining columns are the development periods.

        Returns
        -------
        No return value. The triangle is added to the model.
        """
        # create the triangle
        tri = Triangle.from_csv(id=id, filename=file_path, origin_columns=origin_columns)

        # add the triangle to the model
        self.add_triangle(tri)

    def tri_from_excel(self,
                       filename: str = None,
                       id: str = None,
                       origin_columns: int = 1,
                       sheet_name: str = None,
                       sheet_range: Optional[str] = None,
                       ) -> None:
        """
        Create a triangle from an excel file. Calls the `from_excel` method of the
        `Triangle` class.

        Parameters
        ----------
        filename : str, optional
            The path to the excel file. The default is None, which will
            prompt the user to select a file with a file dialog.
        id : str, optional
            The id of the triangle. The default is None, which will
            prompt the user to enter an id.
        origin_columns : int, optional
            The number of origin columns in the triangle. The default is 1,
            which means that the first column is the the origin (eg accident
            year) and the remaining columns are the development periods.
        sheet_name : str, optional
            The name of the sheet in the excel file. The default is None,
            which will prompt the user to select a sheet from a list of 
            sheet names loaded from the excel file.
        sheet_range : str, optional
            The range of cells in the sheet to load. The default is None,
            which means the entire sheet will be loaded.

        Returns
        -------
        No return value. The triangle is added to the model.
        """
        # if filename is None, prompt the user to select a file
        if filename is None or filename == '':
            filename = filedialog.askopenfilename(title='Select a file',
                                                  filetypes=[('Excel files', '*.xlsx *.xls')])
        
        # if id is None, prompt the user to enter an id
        if id is None or id == '':
            print("You must enter a triangle id.")
            id = input('Enter the id of the triangle: ')

        # if sheet_name is None, prompt the user to select a sheet
        if sheet_name is None or sheet_name == '':
            sheet_name = self._select_sheet_from_excel(filename)

        # check that the sheet name is a string
        if not isinstance(sheet_name, str):
            raise TypeError('The sheet name must be a string.')
        
        # check that the sheet name is a valid sheet name
        if sheet_name not in load_workbook(filename).sheetnames:
            print("The sheets in this workbook are:")
            for sht in load_workbook(filename).sheetnames:
                print(f'  {sht}')
            print("")
            raise ValueError(f'The sheet name "{sheet_name}" is not valid.')
        
        # check that the sheet range is a string in the "A1:B2" format
        if sheet_range is not None:
            if not isinstance(sheet_range, str):
                raise TypeError('The sheet range must be passed as a string.')
            if not re.match(r'^[A-Z]+[0-9]+:[A-Z]+[0-9]+$', sheet_range):
                raise ValueError('The sheet range must be in the "A1:B2" format.')

        # create the triangle
        tri = Triangle.from_excel(filename=filename,
                                  id=id,
                                  origin_columns=origin_columns,
                                  sheet_name=sheet_name,
                                  sheet_range=sheet_range)
        
        # add the triangle to the model
        self.add_triangle(tri)
        


    def add_triangle(self, triangle: Triangle) -> None:
        """
        Add a triangle to the model.
        """
        # check that the triangle is a Triangle object
        if not isinstance(triangle, Triangle):
            raise TypeError('The triangle must be a Triangle object.')
        
        # check that the triangle has a unique id
        if triangle.id in self.tri.__dict__.keys():
            raise ValueError(f'The triangle id "{triangle.id}" is not unique.')
        
        # add the triangle to the model
        self.tri.add(triangle)

    # def add_model(self, model: object) -> None:

    def _select_sheet_from_excel(filename : str) -> str:
        """
        Select a sheet from an excel file.
        """
        # load the excel file
        wb = load_workbook(filename=filename)

        # get the sheet names
        sheet_names = wb.sheetnames

        # prompt the user to select a sheet
        print('Select a sheet from the following list:')
        for i, sheet_name in enumerate(sheet_names):
            print(f'{i+1}. {sheet_name}')
        sheet_name = sheet_names[int(input('Enter the number of the sheet: ')) - 1]

        return sheet_name